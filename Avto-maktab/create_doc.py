import openpyxl
from openpyxl.styles import Alignment, Border, Side, Font
import os

def create_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Dalolatnoma"

    # Sarlavha (Row 1)
    ws.merge_cells('A1:E1')
    ws['A1'] = "Қабул қилиш-топшириш далолатномаси"
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A1'].font = Font(bold=True, size=14)

    header_text = '"Ахборот технологияларини ривожлантириш маркази" МЧЖ томонидан Қорақалпоғистон Республикаси Амударё тумани "Рақамли технологияlar ўқув маркази" га юборилган асосий воситалар ва товар-моддий бойликлар РЎЙХАТИ'
    ws.merge_cells('A2:E4')
    ws['A2'] = header_text
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws['A2'].font = Font(bold=True, size=11)

    # Jadval sarlavhasi (Row 5)
    headers = ["№", "Асосий воситалар ва товар-моддий бойликлар номи", "Ўлчов бирлиги", "Хусусияти", "Сони"]
    for col_num, value in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_num, value=value)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Ma'lumotlar (Rows 6-19)
    data = [
        [1, 'Компьютер в комплекте Монитор 22" (hp) клавиатура, мышка', "дона", "Системный блок IMMER Intel® Core™ i3-8100 CPU @ 3.60GHz OЗY: 4 Гб", 3],
        [2, 'Компьютер в комплекте Монитор 22" (artel) клавиатура, мышка', "дона", "Системный блок NANOTECH Intel® Pentium™ GoldG5420 CPU @ 3.80GHz 3.79GHz", 9],
        [3, 'Принтер 3 в 1', "дона", "canon i-SENSYS MF 112", 1],
        [4, 'Маркерлик доска', "дона", "1.4x0.8", 3],
        [5, 'Ардуино', "дона", "", ""], # Rasmdagi kabi bo'sh qoldirildi
        [6, 'Проектор', "дона", "Epson H839B", 1],
        [7, 'Наушник', "дона", "Kubite T-590", 3],
        [8, "Пилот", "дона", "", 7],
        [9, "Баннер", "дона", "", 6],
        [10, "Артел Телевизор", "дона", "", 1],
        [11, "шкаф", "дона", "", 1],
        [12, "Кондиционер Artel", "дона", "", 2],
        [13, "стул", "дона", "", 10],
        [14, "катта стол", "дона", "", 1]
    ]

    for row_idx, row_data in enumerate(data, 6):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal='center' if col_idx != 2 else 'left', vertical='center', wrap_text=True)
            cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Ustun kengliklari
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 45
    ws.column_dimensions['E'].width = 8

    # Qator balandliklarini sozlash
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 20
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 40

    # Imzolar bo'limi (Starting from Row 22)
    start_sign_row = 22
    
    ws.cell(row=start_sign_row, column=1, value='"Рақамли технологияларни" ўқув маркази (IT-Марказ)')
    ws.cell(row=start_sign_row + 1, column=1, value='Қорақалпоғистон Республикаси ҳудудий филиал директори')
    # İsmlar olib tashlandi
    
    ws.cell(row=start_sign_row + 4, column=1, value='"Рақамли технологияларни" ўқув маркази (IT-Марказ)')
    ws.cell(row=start_sign_row + 5, column=1, value='Амударё туман ҳудудий филиал Топширувчи масъул ходим')
    
    ws.cell(row=start_sign_row + 8, column=1, value='"Рақамли технологияларни" ўқув маркази (IT-Марказ)')
    ws.cell(row=start_sign_row + 9, column=1, value='Амударё туман ҳудудий филиал Қабул қилувчи масъул ходим')

    # Faylni saqlash
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "Qabul_qilish_topshirish_dalolatnomasi_yangi.xlsx")
    wb.save(desktop_path)
    print(f"Fayl yaratildi: {desktop_path}")

if __name__ == "__main__":
    create_excel()
