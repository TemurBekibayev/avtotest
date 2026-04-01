import openpyxl
import os

def debug_excel():
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop", "Qabul_qilish_topshirish_dalolatnomasi.xlsx")
    if not os.path.exists(desktop_path):
        print("Fayl topilmadi.")
        return

    wb = openpyxl.load_workbook(desktop_path)
    ws = wb.active
    
    print(f"Max row: {ws.max_row}")
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=5), 1):
        values = [cell.value for cell in row]
        print(f"Row {i}: {values}")

if __name__ == "__main__":
    debug_excel()
